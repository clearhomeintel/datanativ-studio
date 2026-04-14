import os
import json

COLLEGE_DATA_FILE = os.path.join(os.path.dirname(__file__), "college_data_cache.json")
EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                          "attached_assets", "college_data_populated_v6_updated_1776050789287.xlsx")

SUPPLEMENTAL_DATA = {
    "Brown University": {"location": "Providence, RI", "region": "Northeast", "acceptance_rate": 0.05, "tuition": 65146, "enrollment": 7226, "setting": "Urban", "avg_financial_aid": 58000},
    "Columbia University": {"location": "New York, NY", "region": "Northeast", "acceptance_rate": 0.04, "tuition": 65524, "enrollment": 8216, "setting": "Urban", "avg_financial_aid": 61000},
    "Cornell University": {"location": "Ithaca, NY", "region": "Northeast", "acceptance_rate": 0.08, "tuition": 63200, "enrollment": 15105, "setting": "College Town", "avg_financial_aid": 47000},
    "Dartmouth College": {"location": "Hanover, NH", "region": "Northeast", "acceptance_rate": 0.06, "tuition": 62430, "enrollment": 4318, "setting": "Rural", "avg_financial_aid": 57000},
    "Harvard University": {"location": "Cambridge, MA", "region": "Northeast", "acceptance_rate": 0.04, "tuition": 54768, "enrollment": 7153, "setting": "Urban", "avg_financial_aid": 55000},
    "Princeton University": {"location": "Princeton, NJ", "region": "Northeast", "acceptance_rate": 0.04, "tuition": 57410, "enrollment": 5321, "setting": "Suburban", "avg_financial_aid": 60000},
    "University of Pennsylvania": {"location": "Philadelphia, PA", "region": "Northeast", "acceptance_rate": 0.07, "tuition": 63452, "enrollment": 10496, "setting": "Urban", "avg_financial_aid": 51000},
    "Yale University": {"location": "New Haven, CT", "region": "Northeast", "acceptance_rate": 0.05, "tuition": 62250, "enrollment": 6092, "setting": "Urban", "avg_financial_aid": 57000},
    "MIT": {"location": "Cambridge, MA", "region": "Northeast", "acceptance_rate": 0.04, "tuition": 57986, "enrollment": 4638, "setting": "Urban", "avg_financial_aid": 48000},
    "Stanford University": {"location": "Stanford, CA", "region": "West", "acceptance_rate": 0.04, "tuition": 56169, "enrollment": 7761, "setting": "Suburban", "avg_financial_aid": 52000},
    "Duke University": {"location": "Durham, NC", "region": "South", "acceptance_rate": 0.08, "tuition": 59180, "enrollment": 6751, "setting": "Suburban", "avg_financial_aid": 50000},
    "Northwestern University": {"location": "Evanston, IL", "region": "Midwest", "acceptance_rate": 0.07, "tuition": 60768, "enrollment": 8327, "setting": "Suburban", "avg_financial_aid": 48000},
    "Vanderbilt University": {"location": "Nashville, TN", "region": "South", "acceptance_rate": 0.09, "tuition": 57978, "enrollment": 7113, "setting": "Urban", "avg_financial_aid": 52000},
    "Georgetown University": {"location": "Washington, DC", "region": "South", "acceptance_rate": 0.12, "tuition": 59787, "enrollment": 7518, "setting": "Urban", "avg_financial_aid": 35000},
    "Notre Dame": {"location": "Notre Dame, IN", "region": "Midwest", "acceptance_rate": 0.13, "tuition": 58843, "enrollment": 8863, "setting": "Suburban", "avg_financial_aid": 42000},
    "University of Notre Dame": {"location": "Notre Dame, IN", "region": "Midwest", "acceptance_rate": 0.13, "tuition": 58843, "enrollment": 8863, "setting": "Suburban", "avg_financial_aid": 42000},
    "Rice University": {"location": "Houston, TX", "region": "South", "acceptance_rate": 0.09, "tuition": 52895, "enrollment": 4240, "setting": "Urban", "avg_financial_aid": 45000},
    "Emory University": {"location": "Atlanta, GA", "region": "South", "acceptance_rate": 0.13, "tuition": 57948, "enrollment": 7013, "setting": "Suburban", "avg_financial_aid": 43000},
    "University of Michigan": {"location": "Ann Arbor, MI", "region": "Midwest", "acceptance_rate": 0.20, "tuition": 47476, "enrollment": 31329, "setting": "College Town", "avg_financial_aid": 22000},
    "UCLA": {"location": "Los Angeles, CA", "region": "West", "acceptance_rate": 0.14, "tuition": 42994, "enrollment": 31636, "setting": "Urban", "avg_financial_aid": 24000},
    "UC Berkeley": {"location": "Berkeley, CA", "region": "West", "acceptance_rate": 0.16, "tuition": 43980, "enrollment": 31781, "setting": "Urban", "avg_financial_aid": 23000},
    "Tufts University": {"location": "Medford, MA", "region": "Northeast", "acceptance_rate": 0.11, "tuition": 62374, "enrollment": 6588, "setting": "Suburban", "avg_financial_aid": 42000},
    "Boston University": {"location": "Boston, MA", "region": "Northeast", "acceptance_rate": 0.19, "tuition": 57918, "enrollment": 17588, "setting": "Urban", "avg_financial_aid": 36000},
    "NYU": {"location": "New York, NY", "region": "Northeast", "acceptance_rate": 0.21, "tuition": 56500, "enrollment": 26733, "setting": "Urban", "avg_financial_aid": 27000},
    "University of Southern California": {"location": "Los Angeles, CA", "region": "West", "acceptance_rate": 0.13, "tuition": 62468, "enrollment": 21000, "setting": "Urban", "avg_financial_aid": 35000},
    "Georgia Tech": {"location": "Atlanta, GA", "region": "South", "acceptance_rate": 0.17, "tuition": 32876, "enrollment": 16527, "setting": "Urban", "avg_financial_aid": 19000},
    "University of Virginia": {"location": "Charlottesville, VA", "region": "South", "acceptance_rate": 0.21, "tuition": 52948, "enrollment": 16777, "setting": "College Town", "avg_financial_aid": 20000},
    "University of North Carolina": {"location": "Chapel Hill, NC", "region": "South", "acceptance_rate": 0.23, "tuition": 36776, "enrollment": 19117, "setting": "College Town", "avg_financial_aid": 18000},
    "Williams College": {"location": "Williamstown, MA", "region": "Northeast", "acceptance_rate": 0.09, "tuition": 60590, "enrollment": 2052, "setting": "Rural", "avg_financial_aid": 57000},
    "Amherst College": {"location": "Amherst, MA", "region": "Northeast", "acceptance_rate": 0.11, "tuition": 62830, "enrollment": 1849, "setting": "College Town", "avg_financial_aid": 58000},
    "Pomona College": {"location": "Claremont, CA", "region": "West", "acceptance_rate": 0.08, "tuition": 57476, "enrollment": 1724, "setting": "Suburban", "avg_financial_aid": 54000},
    "Swarthmore College": {"location": "Swarthmore, PA", "region": "Northeast", "acceptance_rate": 0.09, "tuition": 58950, "enrollment": 1663, "setting": "Suburban", "avg_financial_aid": 55000},
    "Middlebury College": {"location": "Middlebury, VT", "region": "Northeast", "acceptance_rate": 0.16, "tuition": 59290, "enrollment": 2700, "setting": "Rural", "avg_financial_aid": 49000},
    "Colby College": {"location": "Waterville, ME", "region": "Northeast", "acceptance_rate": 0.10, "tuition": 62220, "enrollment": 2000, "setting": "Rural", "avg_financial_aid": 50000},
    "Carleton College": {"location": "Northfield, MN", "region": "Midwest", "acceptance_rate": 0.20, "tuition": 58000, "enrollment": 2105, "setting": "Rural", "avg_financial_aid": 48000},
    "Haverford College": {"location": "Haverford, PA", "region": "Northeast", "acceptance_rate": 0.15, "tuition": 58000, "enrollment": 1400, "setting": "Suburban", "avg_financial_aid": 51000},
    "Bowdoin College": {"location": "Brunswick, ME", "region": "Northeast", "acceptance_rate": 0.09, "tuition": 60950, "enrollment": 1858, "setting": "College Town", "avg_financial_aid": 55000},
    "Davidson College": {"location": "Davidson, NC", "region": "South", "acceptance_rate": 0.20, "tuition": 57228, "enrollment": 1900, "setting": "College Town", "avg_financial_aid": 43000},
    "Hamilton College": {"location": "Clinton, NY", "region": "Northeast", "acceptance_rate": 0.13, "tuition": 59760, "enrollment": 2000, "setting": "Rural", "avg_financial_aid": 52000},
    "Colgate University": {"location": "Hamilton, NY", "region": "Northeast", "acceptance_rate": 0.22, "tuition": 59545, "enrollment": 3100, "setting": "Rural", "avg_financial_aid": 48000},
    "Bucknell University": {"location": "Lewisburg, PA", "region": "Northeast", "acceptance_rate": 0.37, "tuition": 57906, "enrollment": 3650, "setting": "Rural", "avg_financial_aid": 38000},
    "Lehigh University": {"location": "Bethlehem, PA", "region": "Northeast", "acceptance_rate": 0.38, "tuition": 57680, "enrollment": 5700, "setting": "Suburban", "avg_financial_aid": 33000},
}

def load_college_data():
    if os.path.exists(COLLEGE_DATA_FILE):
        with open(COLLEGE_DATA_FILE, "r") as f:
            return json.load(f)
    return _parse_excel()

def _safe_float(val, default=None):
    try:
        if val is None or str(val).strip() in ["", "N/A", "nan", "None"]:
            return default
        return float(str(val).replace("%","").replace(",","").strip())
    except Exception:
        return default

def _safe_int(val, default=None):
    v = _safe_float(val, default)
    return int(v) if v is not None else default

def _parse_excel():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        parsed = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            record = dict(zip(headers, row))
            name = str(record.get("University_Name","")).strip()
            if not name or name == "None":
                continue

            sat_math_75 = _safe_int(record.get("SAT_Math_75th"))
            sat_ebrw_75 = _safe_int(record.get("SAT_EBRW_75th"))
            sat_math_50 = _safe_int(record.get("SAT_Math_50th"))
            sat_ebrw_50 = _safe_int(record.get("SAT_EBRW_50th"))
            act_75 = _safe_int(record.get("ACT_Composite_75th"))
            act_50 = _safe_int(record.get("ACT_Composite_50th"))

            sat_combined = None
            if sat_math_50 and sat_ebrw_50:
                sat_combined = sat_math_50 + sat_ebrw_50
            elif sat_math_75 and sat_ebrw_75:
                sat_combined = sat_math_75 + sat_ebrw_75

            suppl = SUPPLEMENTAL_DATA.get(name, {})

            college = {
                "name": name,
                "type": str(record.get("Type","")).strip(),
                "location": suppl.get("location",""),
                "region": suppl.get("region",""),
                "sat_range": sat_combined or 1350,
                "act_range": act_50 or act_75 or 30,
                "acceptance_rate": suppl.get("acceptance_rate", 0.20),
                "avg_financial_aid": suppl.get("avg_financial_aid", 30000),
                "tuition": suppl.get("tuition", 55000),
                "enrollment": suppl.get("enrollment", 5000),
                "setting": suppl.get("setting","Suburban"),
                "cds_factors": {
                    "rigor": _safe_float(record.get("C7_Rigor"), 0),
                    "gpa": _safe_float(record.get("C7_GPA"), 0),
                    "test_scores": _safe_float(record.get("C7_TestScores"), 0),
                    "essay": _safe_float(record.get("C7_Essay"), 0),
                    "extracurriculars": _safe_float(record.get("C7_Extracurriculars"), 0),
                    "recommendations": _safe_float(record.get("C7_Recommendation"), 0),
                    "interview": _safe_float(record.get("C7_Interview"), 0),
                }
            }
            parsed.append(college)

        with open(COLLEGE_DATA_FILE, "w") as f:
            json.dump(parsed, f, indent=2, default=str)

        return parsed

    except Exception as e:
        print(f"Excel parse error: {e}")
        return _get_fallback_colleges()

def _get_fallback_colleges():
    return list({k: v for k, v in SUPPLEMENTAL_DATA.items()}.items())

COLLEGE_DATABASE = None

def get_colleges():
    global COLLEGE_DATABASE
    if COLLEGE_DATABASE is None:
        COLLEGE_DATABASE = load_college_data()
    return COLLEGE_DATABASE
